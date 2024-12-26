import re
import openpyxl

wb = openpyxl.load_workbook('energomera.xlsx')
ws = wb.active

pattern1 = re.compile(r'\d{6}_\d{3}')

pattern2 = re.compile(r'\d{6}_\d{3}_\d{2}')

for row in ws.iter_rows():
    for cell in row:
        match1 = pattern1.search(str(cell.value))

        if match1:
            new_value = match1.group().replace('_', '.')

            next_column = cell.column + 2
            next_row = cell.row
            target_cell = ws.cell(next_row, next_column)
            target_cell.value = new_value

        match2 = pattern2.search(str(cell.value))

        if match2:
            new_value = match2.group().replace('_', '.', 1).replace('_', '-')

            next_column = cell.column + 2
            next_row = cell.row
            target_cell = ws.cell(next_row, next_column)
            target_cell.value = new_value

wb.save('updated_file.xlsx')